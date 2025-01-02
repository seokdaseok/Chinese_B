import random

from openpyxl import load_workbook

normal_words_directory = "Words/normal_word_bank.xlsx"

listening_directory = "Words/listening_word_bank.xlsx"

the_current_word_index = 1
word_display = ["Pin Yin", "詞語", "Lorem ipsum dolor sit amet, consectetur adipiscing elit, sed do eiusmod tempor incididunt ut labore et dolore magna aliqua.", 0]

prof_change_amount = 10

def set_directory_normal(the_string):
    global normal_words_directory
    normal_words_directory = the_string

def set_directory_listening(the_string):
    global listening_directory
    listening_directory = the_string

def add_word(word, pinyin, category, definition, topic):
    global normal_words_directory

    ##new_row = {'Word': word, 'Pinyin': pinyin, 'Category': category, 'Definition': definition, 'Profficiency': 0, 'time_since_encounter': 0}

    wb = load_workbook(normal_words_directory)
    ws = wb.active

    word_cell = ws.cell(row=ws.max_row+1, column=1)
    word_cell.value = word

    pinyin_cell = ws.cell(row=ws.max_row, column=2)
    pinyin_cell.value = pinyin

    category_cell = ws.cell(row=ws.max_row, column=3)
    category_cell.value = category

    definition_cell = ws.cell(row=ws.max_row, column=4)
    definition_cell.value = definition

    profficiency_cell = ws.cell(row=ws.max_row, column=5)
    profficiency_cell.value = 0 

    time_since_encounter_cell = ws.cell(row=ws.max_row, column=6)
    time_since_encounter_cell.value = 0 

    topic_cell = ws.cell(row=ws.max_row, column=7)
    topic_cell.value = topic

    wb.save(normal_words_directory)

test_low_profficiency = False

def next_word(word_type, word_topic):
    global test_low_profficiency
    global the_current_word_index
    global word_display

    # profficiency recursion function

    prof_distribution = [0,0,0,0,0,0,0,0,0,0,0,0,
                         20,20,20,20,20,20,
                         40,40,40,
                         60,60,
                         80]

    def prof_recursion(indexes, prof):

        indexes_of_prof = []

        for i in indexes:
            _prof = ws.cell(row=i, column=5).value

            if _prof <= prof:
                indexes_of_prof.append(i)
                
        if(len(indexes_of_prof) == 0):
            new_prof = random.choice(prof_distribution)
            indexes_of_prof = prof_recursion(indexes, new_prof)

        return indexes_of_prof
        
    if(word_type =="listening"):
        wb = load_workbook(listening_directory)
        ws = wb.active

        number_of_rows = ws.max_row

        random_word_index = random.randint(2, number_of_rows)

        the_word = ws.cell(row=random_word_index, column=1).value
        the_pinyin = ws.cell(row=random_word_index, column=2).value
        the_definition = ws.cell(row=random_word_index, column=4).value
        the_prof = ws.cell(row=random_word_index, column=5).value
        the_example = ws.cell(row=random_word_index, column=8).value

        word_display = [the_pinyin, the_word, the_definition, the_prof, the_example]

        ws.cell(row=random_word_index, column=6).value = ws.cell(row=random_word_index, column=6).value + 1

        the_current_word_index = random_word_index

    else:

        wb = load_workbook(normal_words_directory)
        ws = wb.active

        number_of_rows = ws.max_row

        ## for any topic

        if(word_topic==""):

            indexes_of_type = []

            for i in range(2, number_of_rows + 1):

                _type = ws.cell(row=i, column=3).value

                if(_type == word_type):
                    indexes_of_type.append(i)

            
            if(len(indexes_of_type) == 0):
                
                for i in range(2, number_of_rows + 1):
                    indexes_of_type.append(i)

            chosen_prof = random.choice(prof_distribution)

            prof_indexes = prof_recursion(indexes_of_type, chosen_prof)

            random_word_index = random.choice(prof_indexes)

            the_word = ws.cell(row=random_word_index, column=1).value
            the_pinyin = ws.cell(row=random_word_index, column=2).value
            the_definition = ws.cell(row=random_word_index, column=4).value
            the_prof = ws.cell(row=random_word_index, column=5).value
            the_example = ws.cell(row=random_word_index, column=8).value

            word_display = [the_pinyin, the_word, the_definition, the_prof, the_example]

            ws.cell(row=random_word_index, column=6).value = ws.cell(row=random_word_index, column=6).value + 1

            the_current_word_index = random_word_index
        
        else:

            indexes_of_type = []

            for i in range(2, number_of_rows + 1):

                _type = ws.cell(row=i, column=3).value

                if(_type == word_type):
                    indexes_of_type.append(i)

            
            if(len(indexes_of_type) == 0):
                
                for i in range(2, number_of_rows + 1):
                    indexes_of_type.append(i)


            indexes_of_topic = []

            for j in indexes_of_type:
                _topic = str(ws.cell(row=j, column=7).value)

                if(_topic == str(word_topic)):
                    indexes_of_topic.append(j)
            
            if(len(indexes_of_topic) == 0):

                for j in indexes_of_type:
                    indexes_of_topic.append(j)

            chosen_prof = random.choice(prof_distribution)

            prof_indexes = prof_recursion(indexes_of_topic, chosen_prof)

            random_word_index = random.choice(prof_indexes)

            the_word = ws.cell(row=random_word_index, column=1).value
            the_pinyin = ws.cell(row=random_word_index, column=2).value
            the_definition = ws.cell(row=random_word_index, column=4).value
            the_prof = ws.cell(row=random_word_index, column=5).value
            the_example = ws.cell(row=random_word_index, column=8).value

            word_display = [the_pinyin, the_word, the_definition, the_prof, the_example]

            ws.cell(row=random_word_index, column=6).value = ws.cell(row=random_word_index, column=6).value + 1

            the_current_word_index=random_word_index

def word_wrong(word_type):
    global the_current_word_index

    if(word_type == 'listening'):
        wb = load_workbook(listening_directory)
        ws = wb.active

        if (ws.cell(row=the_current_word_index, column=5).value >= prof_change_amount):
            ws.cell(row=the_current_word_index, column=5).value = ws.cell(row=the_current_word_index, column=5).value - prof_change_amount

        wb.save(listening_directory)

        word_display[3] = ws.cell(row=the_current_word_index, column=5).value
    else:
        wb = load_workbook(normal_words_directory)
        ws = wb.active

        if (ws.cell(row=the_current_word_index, column=5).value >= prof_change_amount):
            ws.cell(row=the_current_word_index, column=5).value = ws.cell(row=the_current_word_index, column=5).value - prof_change_amount

        wb.save(normal_words_directory)

        word_display[3] = ws.cell(row=the_current_word_index, column=5).value

def word_right(word_type):
    global the_current_word_index

    if(word_type == 'listening'):
        wb = load_workbook(listening_directory)
        ws = wb.active

        if (ws.cell(row=the_current_word_index, column=5).value <= 100 - prof_change_amount):
            ws.cell(row=the_current_word_index, column=5).value = ws.cell(row=the_current_word_index, column=5).value + prof_change_amount

        wb.save(listening_directory)

        word_display[3] = ws.cell(row=the_current_word_index, column=5).value
    else:
        wb = load_workbook(normal_words_directory)
        ws = wb.active

        #print(the_current_word_index)


        if (ws.cell(row=the_current_word_index, column=5).value <= 100 - prof_change_amount):
            ws.cell(row=the_current_word_index, column=5).value = ws.cell(row=the_current_word_index, column=5).value + prof_change_amount

        wb.save(normal_words_directory)

        word_display[3] = ws.cell(row=the_current_word_index, column=5).value








    
